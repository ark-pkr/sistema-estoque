from io import BytesIO

from django.db.models import F
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.response import Response
from rest_framework.views import APIView

from produtos.models import Produto
from usuarios.permissions import IsAuthenticatedAndActive


class RelatorioEstoqueView(APIView):
    permission_classes = [IsAuthenticatedAndActive]

    def get_queryset(self, request):
        queryset = Produto.objects.select_related("categoria", "fornecedor").all().order_by("nome")

        status_filtro = request.GET.get("status")
        categoria_id = request.GET.get("categoria")
        fornecedor_id = request.GET.get("fornecedor")
        busca = request.GET.get("busca")

        if status_filtro == "estoque_baixo":
            queryset = queryset.filter(
                estoque_atual__gt=0,
                estoque_atual__lte=F("estoque_minimo"),
            )
        elif status_filtro == "sem_estoque":
            queryset = queryset.filter(estoque_atual__lte=0)
        elif status_filtro == "normal":
            queryset = queryset.filter(estoque_atual__gt=F("estoque_minimo"))

        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)

        if fornecedor_id:
            queryset = queryset.filter(fornecedor_id=fornecedor_id)

        if busca:
            queryset = queryset.filter(nome__icontains=busca)

        return queryset

    def serialize_data(self, queryset):
        data = []
        for item in queryset:
            estoque_atual = float(getattr(item, "estoque_atual", 0) or 0)
            estoque_minimo = float(getattr(item, "estoque_minimo", 0) or 0)

            if estoque_atual <= 0:
                status = "Sem estoque"
            elif estoque_atual <= estoque_minimo:
                status = "Estoque baixo"
            else:
                status = "Normal"

            data.append(
                {
                    "codigo": getattr(item, "codigo", "") or "-",
                    "nome": getattr(item, "nome", "") or "-",
                    "categoria": getattr(getattr(item, "categoria", None), "nome", "-"),
                    "fornecedor": getattr(getattr(item, "fornecedor", None), "nome", "-"),
                    "marca": getattr(item, "marca", "") or "-",
                    "unidade": getattr(item, "unidade", "") or "-",
                    "preco_venda": float(getattr(item, "preco_venda", 0) or 0),
                    "estoque_atual": estoque_atual,
                    "estoque_minimo": estoque_minimo,
                    "status": status,
                }
            )
        return data

    def get(self, request):
        formato = request.GET.get("formato", "json").lower()
        queryset = self.get_queryset(request)
        data = self.serialize_data(queryset)

        if formato == "pdf":
            return self.exportar_pdf(data)

        if formato == "xlsx":
            return self.exportar_excel(data)

        return Response(
            {
                "total": len(data),
                "results": data,
            }
        )

    def exportar_pdf(self, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Relatório de Estoque", styles["Title"]))
        elements.append(Spacer(1, 12))

        tabela_dados = [
            [
                "Código",
                "Produto",
                "Categoria",
                "Fornecedor",
                "Marca",
                "Un.",
                "Preço",
                "Atual",
                "Mínimo",
                "Status",
            ]
        ]

        for item in data:
            tabela_dados.append(
                [
                    item["codigo"],
                    item["nome"],
                    item["categoria"],
                    item["fornecedor"],
                    item["marca"],
                    item["unidade"],
                    f'R$ {item["preco_venda"]:.2f}',
                    f'{item["estoque_atual"]:.2f}',
                    f'{item["estoque_minimo"]:.2f}',
                    item["status"],
                ]
            )

        tabela = Table(tabela_dados, repeatRows=1)
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )

        elements.append(tabela)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="relatorio_estoque.pdf"'
        return response

    def exportar_excel(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Estoque"

        headers = [
            "Código",
            "Produto",
            "Categoria",
            "Fornecedor",
            "Marca",
            "Unidade",
            "Preço de Venda",
            "Estoque Atual",
            "Estoque Mínimo",
            "Status",
        ]
        ws.append(headers)

        fill = PatternFill("solid", fgColor="0F172A")
        font = Font(color="FFFFFF", bold=True)

        for col_index, _header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index)
            cell.fill = fill
            cell.font = font

        for item in data:
            ws.append(
                [
                    item["codigo"],
                    item["nome"],
                    item["categoria"],
                    item["fornecedor"],
                    item["marca"],
                    item["unidade"],
                    item["preco_venda"],
                    item["estoque_atual"],
                    item["estoque_minimo"],
                    item["status"],
                ]
            )

        widths = [18, 32, 22, 24, 18, 12, 16, 16, 16, 18]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="relatorio_estoque.xlsx"'
        return response